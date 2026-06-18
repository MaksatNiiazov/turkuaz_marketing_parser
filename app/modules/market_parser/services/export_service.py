from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.modules.market_parser.repositories.product_repo import ProductRepository
from app.modules.market_parser.repositories.snapshot_repo import SnapshotRepository
from app.modules.market_parser.services.stats_service import StatsService, effective_price


class ExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.products = ProductRepository(db)
        self.snapshots = SnapshotRepository(db)
        self.stats = StatsService(db)

    def products_xlsx(
        self,
        source_id: int | None = None,
        category_id: int | None = None,
        name: str | None = None,
        sku: str | None = None,
        has_discount: bool | None = None,
        is_available: bool | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
        run_id: int | None = None,
    ) -> BytesIO:
        products = self.products.list(source_id=source_id, category_id=category_id, name=name, sku=sku)
        if run_id is None and (from_date is not None or to_date is not None):
            return self._period_products_xlsx(products, from_date, to_date, has_discount, is_available)
        product_snapshots = self._product_snapshots_for_export(products, from_date, to_date, run_id)
        wb = self._base_workbook()
        ws = wb["Товары"]
        product_rows = []
        self._write_header(
            ws,
            [
                "source",
                "sku",
                "external_id",
                "name",
                "title",
                "parent_category",
                "category",
                "price",
                "discount_price",
                "media",
                "product_url",
                "availability",
                "collected_at",
            ],
        )
        for product, snapshot in product_snapshots:
            if has_discount is not None:
                discounted = bool(
                    snapshot and (snapshot.discount_price is not None or snapshot.discount_percent is not None)
                )
                if discounted != has_discount:
                    continue
            if is_available is not None and (snapshot is None or snapshot.is_available != is_available):
                continue
            row = self._product_export_row(product, snapshot)
            product_rows.append(row)
            ws.append(row)
        self._category_sheets(wb, product_rows)
        self._summary_sheet(wb)
        return save_workbook(wb)

    def _product_snapshots_for_export(
        self,
        products: list,
        from_date: date | None,
        to_date: date | None,
        run_id: int | None,
    ) -> list[tuple]:
        product_by_id = {product.id: product for product in products}
        if run_id is not None:
            snapshots = self.snapshots.list_for_run(run_id, from_date, to_date)
            return [
                (product_by_id[snapshot.product_id], snapshot)
                for snapshot in snapshots
                if snapshot.product_id in product_by_id
            ]
        if from_date is not None or to_date is not None:
            snapshots = self.snapshots.list_for_product_ids(list(product_by_id), from_date, to_date)
            return [
                (product_by_id[snapshot.product_id], snapshot)
                for snapshot in snapshots
                if snapshot.product_id in product_by_id
            ]
        latest = {
            snapshot.product_id: snapshot
            for snapshot in self.snapshots.latest_by_product_ids(list(product_by_id))
        }
        return [(product, latest.get(product.id)) for product in products]

    def _period_products_xlsx(
        self,
        products: list,
        from_date: date | None,
        to_date: date | None,
        has_discount: bool | None,
        is_available: bool | None,
    ) -> BytesIO:
        product_by_id = {product.id: product for product in products}
        snapshots = self.snapshots.list_for_product_ids(list(product_by_id), from_date, to_date)
        daily_snapshots = choose_daily_snapshots(snapshots)
        days = sorted({day for _, day in daily_snapshots})

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Товары")
        headers = [
            "source",
            "sku",
            "external_id",
            "name",
            "title",
            "parent_category",
            "category",
            "media",
            "product_url",
        ]
        for day in days:
            label = day.isoformat()
            headers.extend(
                [
                    f"{label} price",
                    f"{label} discount_price",
                    f"{label} discount_percent",
                ]
            )
        ws.append(headers)

        snapshots_by_product: dict[int, dict[date, object]] = defaultdict(dict)
        for (product_id, day), snapshot in daily_snapshots.items():
            snapshots_by_product[product_id][day] = snapshot

        for product in products:
            product_daily = snapshots_by_product.get(product.id, {})
            if not product_daily:
                continue
            if not product_matches_filters(product_daily.values(), has_discount, is_available):
                continue
            parent_name, category_name = category_names(product)
            row = [
                product.source.code,
                product.sku,
                product.external_sku,
                product.name,
                product.unit,
                parent_name,
                category_name,
                product.image_url,
                product.product_url,
            ]
            for day in days:
                snapshot = product_daily.get(day)
                row.extend(
                    [
                        snapshot.price if snapshot else None,
                        snapshot.discount_price if snapshot else None,
                        snapshot.discount_percent if snapshot else None,
                    ]
                )
            ws.append(row)

        summary = wb.create_sheet("Свод")
        summary.append(["Метрика", "Значение"])
        summary.append(["Источник", "Globus Online"])
        summary.append(["Формат", "Одна строка на товар, даты в колонках"])
        summary.append(["Дней в выгрузке", len(days)])
        return save_workbook_fast(wb)

    def stats_xlsx(
        self, from_date: date | None = None, to_date: date | None = None, category_id: int | None = None
    ) -> BytesIO:
        wb = self._base_workbook()
        products = self.products.list(category_id=category_id)
        product_ids = [product.id for product in products]
        latest = {s.product_id: s for s in self.snapshots.latest_by_product_ids(product_ids)}

        ws_products = wb["Товары"]
        product_rows = []
        self._write_header(
            ws_products,
            [
                "sku",
                "external_id",
                "name",
                "title",
                "parent_category",
                "category",
                "current_price",
                "discount_price",
                "media",
                "availability",
            ],
        )
        for product in products:
            snapshot = latest.get(product.id)
            parent_name, category_name = category_names(product)
            row = [
                product.sku,
                product.external_sku,
                product.name,
                product.unit,
                parent_name,
                category_name,
                effective_price(snapshot) if snapshot else None,
                snapshot.discount_price if snapshot else None,
                product.image_url,
                availability_text(snapshot.is_available if snapshot else None),
            ]
            product_rows.append(row)
            ws_products.append(row)

        ws_prices = wb["Цены"]
        self._write_header(
            ws_prices,
            ["date", "sku", "external_id", "name", "price", "discount_price", "effective_price", "availability"],
        )
        ws_discounts = wb["Скидки"]
        self._write_header(
            ws_discounts,
            ["date", "sku", "external_id", "name", "price", "discount_price", "discount_percent", "availability"],
        )
        for product in products:
            for snapshot in self.snapshots.list_for_product(product.id, from_date, to_date):
                ws_prices.append(
                    [
                        snapshot.collected_at,
                        product.sku,
                        product.external_sku,
                        product.name,
                        snapshot.price,
                        snapshot.discount_price,
                        effective_price(snapshot),
                        availability_text(snapshot.is_available),
                    ]
                )
                if snapshot.discount_price is not None or snapshot.discount_percent is not None:
                    ws_discounts.append(
                        [
                            snapshot.collected_at,
                            product.sku,
                            product.external_sku,
                            product.name,
                            snapshot.price,
                            snapshot.discount_price,
                            snapshot.discount_percent,
                            availability_text(snapshot.is_available),
                        ]
                    )

        ws_changes = wb["Изменения"]
        self._write_header(ws_changes, ["product_id", "name", "first_price", "last_price", "change_percent"])
        if category_id is not None:
            for item in self.stats.price_changes(from_date, to_date, category_id=category_id):
                ws_changes.append(
                    [item.product_id, item.name, item.first_price, item.last_price, item.change_percent]
                )

        self._category_sheets(
            wb,
            product_rows,
            headers=[
                "sku",
                "external_id",
                "name",
                "title",
                "parent_category",
                "category",
                "current_price",
                "discount_price",
                "media",
                "availability",
            ],
            parent_index=4,
        )
        self._summary_sheet(wb)
        return save_workbook(wb)

    def _product_export_row(self, product, snapshot) -> list:
        parent_name, category_name = category_names(product)
        return [
            product.source.code,
            product.sku,
            product.external_sku,
            product.name,
            product.unit,
            parent_name,
            category_name,
            snapshot.price if snapshot else None,
            snapshot.discount_price if snapshot else None,
            product.image_url,
            product.product_url,
            availability_text(snapshot.is_available if snapshot else None),
            snapshot.collected_at if snapshot else None,
        ]

    def _category_sheets(
        self,
        wb: Workbook,
        rows: list[list],
        headers: list[str] | None = None,
        parent_index: int = 5,
    ) -> None:
        if headers is None:
            headers = [
                "source",
                "sku",
                "external_id",
                "name",
                "title",
                "parent_category",
                "category",
                "price",
                "discount_price",
                "media",
                "product_url",
                "availability",
                "collected_at",
            ]
        grouped: dict[str, list[list]] = {}
        for row in rows:
            parent_name = row[parent_index] or "Без категории"
            grouped.setdefault(str(parent_name), []).append(row)
        for parent_name, parent_rows in sorted(grouped.items()):
            title = unique_sheet_title(wb, parent_name)
            ws = wb.create_sheet(title)
            self._write_header(ws, headers)
            for row in parent_rows:
                ws.append(row)

    def _base_workbook(self) -> Workbook:
        wb = Workbook()
        wb.active.title = "Товары"
        for title in ["Цены", "Скидки", "Изменения", "Свод"]:
            wb.create_sheet(title)
        return wb

    def _summary_sheet(self, wb: Workbook) -> None:
        ws = wb["Свод"]
        self._write_header(ws, ["Метрика", "Значение"])
        ws.append(["Источник", "Globus Online"])
        ws.append(["Совместимые поля", "sku, external_id, name, title, price, discount_price, media"])

    @staticmethod
    def _write_header(ws, values: list[str]) -> None:
        ws.append(values)
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        ws.freeze_panes = "A2"


def save_workbook(wb: Workbook) -> BytesIO:
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def save_workbook_fast(wb: Workbook) -> BytesIO:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def category_names(product) -> tuple[str | None, str | None]:
    category = product.category
    if category is None:
        return None, None
    if category.parent is not None:
        return category.parent.name, category.name
    return category.name, category.name


def availability_text(value: bool | None) -> str:
    if value is True:
        return "доступно"
    if value is False:
        return "нет в наличии"
    return "неизвестно"


def choose_daily_snapshots(snapshots: list) -> dict[tuple[int, date], object]:
    run_counts_by_day: dict[date, dict[int | None, int]] = defaultdict(lambda: defaultdict(int))
    for snapshot in snapshots:
        run_counts_by_day[snapshot.collected_at.date()][snapshot.run_id] += 1

    selected_run_by_day = {
        day: max(run_counts.items(), key=lambda item: (item[1], item[0] or 0))[0]
        for day, run_counts in run_counts_by_day.items()
    }

    selected: dict[tuple[int, date], object] = {}
    for snapshot in snapshots:
        day = snapshot.collected_at.date()
        if snapshot.run_id != selected_run_by_day[day]:
            continue
        key = (snapshot.product_id, day)
        current = selected.get(key)
        if current is None or snapshot_quality_key(snapshot) > snapshot_quality_key(current):
            selected[key] = snapshot
    return selected


def snapshot_quality_key(snapshot) -> tuple[int, object]:
    score = 0
    if snapshot.price is not None:
        score += 4
    if snapshot.discount_price is not None:
        score += 2
    if snapshot.discount_percent is not None:
        score += 1
    return score, snapshot.collected_at


def product_matches_filters(snapshots, has_discount: bool | None, is_available: bool | None) -> bool:
    rows = list(snapshots)
    if has_discount is not None:
        matches_discount = any(
            (snapshot.discount_price is not None or snapshot.discount_percent is not None) == has_discount
            for snapshot in rows
        )
        if not matches_discount:
            return False
    if is_available is not None and not any(snapshot.is_available == is_available for snapshot in rows):
        return False
    return True


def unique_sheet_title(wb: Workbook, value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", " ", value).strip() or "Категория"
    title = title[:31]
    if title not in wb.sheetnames:
        return title
    base = title[:28]
    index = 2
    while f"{base} {index}" in wb.sheetnames:
        index += 1
    return f"{base} {index}"
