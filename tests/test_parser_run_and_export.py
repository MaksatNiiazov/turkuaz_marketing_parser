import asyncio
from datetime import datetime, timezone
from decimal import Decimal

import pytest
import httpx
from openpyxl import load_workbook

from app.modules.market_parser.models.entities import ParserCategory, ParserRun, ParserRunCategory, ParserSource
from app.modules.market_parser.repositories.product_repo import ProductRepository
from app.modules.market_parser.schemas.run import RunCreate
from app.modules.market_parser.services.export_service import ExportService
from app.modules.market_parser.services.globus_parser import ParsedProduct
from app.modules.market_parser.services.parser_service import ParserService
from app.modules.market_parser.services.parser_service import summarize_parser_error
from app.modules.market_parser.services.run_control import request_run_stop
from app.modules.market_parser.services.run_recovery import recover_interrupted_runs
from app.modules.market_parser.services.snapshot_service import SnapshotService


class FakeParser:
    async def fetch_categories(self):
        return []

    async def fetch_products_by_category(self, category):
        if category.name == "Bad":
            raise RuntimeError("category failed")
        return [
            ParsedProduct(
                source_code="globus",
                external_sku=f"external-{category.id}",
                sku=f"sku-{category.id}",
                name=f"Product {category.id}",
                unit="шт.",
                category_name=category.name,
                category_url=category.url,
                price=Decimal("100.00"),
                discount_price=None,
                discount_percent=None,
                image_url=None,
                product_url=f"https://globus-online.kg/ru-kg/good/sku-{category.id}",
                is_available=True,
                raw_data={"id": category.id},
            )
        ]


class SlowParser:
    async def fetch_categories(self):
        return []

    async def fetch_products_by_category(self, category):
        await asyncio.sleep(30)
        return []


def seed_source_categories(db_session):
    source = ParserSource(name="Globus Online", code="globus", base_url="https://globus-online.kg/ru-kg")
    db_session.add(source)
    db_session.flush()
    good = ParserCategory(source_id=source.id, external_id="good", name="Good", url="https://example/good")
    bad = ParserCategory(source_id=source.id, external_id="bad", name="Bad", url="https://example/bad")
    db_session.add_all([good, bad])
    db_session.flush()
    db_session.commit()
    return source, good, bad


def test_summarize_globus_http_error() -> None:
    request = httpx.Request("GET", "https://globus-online.kg/ru-kg/catalog/grocery/category/missing")
    response = httpx.Response(404, request=request)
    error = httpx.HTTPStatusError("Client error", request=request, response=response)

    assert summarize_parser_error(error) == "Globus вернул 404: раздел не найден или больше недоступен"


@pytest.mark.asyncio
async def test_run_keeps_going_when_category_fails(db_session, monkeypatch) -> None:
    source, good, bad = seed_source_categories(db_session)
    monkeypatch.setattr(ParserService, "_parser_for_source", lambda self, code, base_url: FakeParser())

    run = await ParserService(db_session).run_parser(
        RunCreate(source_id=source.id, category_ids=[good.id, bad.id])
    )

    assert run.status == "partial"
    assert run.saved_products == 1
    assert "category failed" in (run.error_message or "")
    assert len(ProductRepository(db_session).list()) == 1


@pytest.mark.asyncio
async def test_run_can_be_stopped(db_session, monkeypatch) -> None:
    source, good, _ = seed_source_categories(db_session)
    monkeypatch.setattr(ParserService, "_parser_for_source", lambda self, code, base_url: SlowParser())
    payload = RunCreate(source_id=source.id, category_ids=[good.id])
    service = ParserService(db_session)
    run = service.create_parser_run(payload)

    task = asyncio.create_task(service.execute_parser_run(payload, run.id))
    for _ in range(10):
        if request_run_stop(run.id):
            break
        await asyncio.sleep(0)
    finished = await asyncio.wait_for(task, timeout=3)

    assert finished.status == "stopped"
    assert finished.processed_categories == 0
    assert "Остановлено пользователем" in (finished.error_message or "")


def test_recover_interrupted_stopping_run(db_session) -> None:
    source, good, _ = seed_source_categories(db_session)
    run = ParserRun(
        source_id=source.id,
        status="stopping",
        started_at=datetime(2026, 6, 4, tzinfo=timezone.utc),
        total_categories=1,
        processed_categories=0,
        total_products=0,
        saved_products=0,
        error_message="Остановка запрошена пользователем",
    )
    db_session.add(run)
    db_session.flush()
    db_session.add(
        ParserRunCategory(
            run_id=run.id,
            category_id=good.id,
            status="running",
            started_at=datetime(2026, 6, 4, tzinfo=timezone.utc),
        )
    )
    db_session.commit()

    recovered = recover_interrupted_runs(db_session)
    db_session.commit()

    assert recovered == 1
    assert run.status == "stopped"
    assert run.finished_at is not None
    assert "Остановлено после перезапуска сервера" in (run.error_message or "")
    assert run.categories[0].status == "stopped"


def test_export_xlsx(db_session) -> None:
    source, _, _ = seed_source_categories(db_session)
    parent = ParserCategory(
        source_id=source.id,
        external_id="home",
        name="Дом и Уют",
        url="https://example/home",
        is_enabled=False,
    )
    db_session.add(parent)
    db_session.flush()
    category = ParserCategory(
        source_id=source.id,
        external_id="kitchen",
        name="Для кухни",
        url="https://example/kitchen",
        parent_id=parent.id,
    )
    db_session.add(category)
    db_session.flush()
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=category.id,
        run_id=None,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="external-1",
            sku="Ц0180022",
            name="Шоколад",
            unit="шт.",
            category_name=category.name,
            category_url=category.url,
            price=Decimal("100.00"),
            discount_price=Decimal("80.00"),
            discount_percent=Decimal("20.00"),
            image_url="https://img",
            product_url="https://globus-online.kg/ru-kg/good/external-1",
            is_available=True,
            raw_data={"id": "external-1", "pigeonId": "Ц0180022"},
        ),
        collected_at=datetime(2026, 5, 29, tzinfo=timezone.utc),
    )
    db_session.commit()

    output = ExportService(db_session).stats_xlsx(category_id=category.id)
    workbook = load_workbook(output)

    assert "Дом и Уют" in workbook.sheetnames
    assert workbook["Товары"]["A1"].value == "sku"
    assert workbook["Товары"]["B1"].value == "external_id"
    assert workbook["Товары"]["A2"].value == "Ц0180022"
    assert workbook["Товары"]["B2"].value == "external-1"
    assert workbook["Товары"]["E2"].value == "Дом и Уют"
    assert workbook["Товары"]["F2"].value == "Для кухни"
    assert workbook["Товары"]["J2"].value == "доступно"
    assert workbook["Скидки"]["D2"].value == "Шоколад"
    assert workbook["Скидки"]["H2"].value == "доступно"


def test_products_export_filters_by_period_and_discount(db_session) -> None:
    source, good, _ = seed_source_categories(db_session)
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=None,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="discounted",
            sku="D-1",
            name="Товар со скидкой",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("120.00"),
            discount_price=Decimal("90.00"),
            discount_percent=Decimal("25.00"),
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/discounted",
            is_available=True,
            raw_data={"id": "discounted"},
        ),
        collected_at=datetime(2026, 6, 10, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=None,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="discounted",
            sku="D-1",
            name="Товар со скидкой",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("110.00"),
            discount_price=Decimal("88.00"),
            discount_percent=Decimal("20.00"),
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/discounted",
            is_available=True,
            raw_data={"id": "discounted"},
        ),
        collected_at=datetime(2026, 6, 10, 12, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=None,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="regular",
            sku="R-1",
            name="Товар без скидки",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("80.00"),
            discount_price=None,
            discount_percent=None,
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/regular",
            is_available=True,
            raw_data={"id": "regular"},
        ),
        collected_at=datetime(2026, 6, 10, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=None,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="outside-period",
            sku="O-1",
            name="Вне периода",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("70.00"),
            discount_price=Decimal("60.00"),
            discount_percent=Decimal("14.29"),
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/outside-period",
            is_available=True,
            raw_data={"id": "outside-period"},
        ),
        collected_at=datetime(2026, 6, 11, tzinfo=timezone.utc),
    )
    db_session.commit()

    discounted_output = ExportService(db_session).products_xlsx(
        category_id=good.id,
        has_discount=True,
        from_date=datetime(2026, 6, 10).date(),
        to_date=datetime(2026, 6, 10).date(),
    )
    discounted_sheet = load_workbook(discounted_output)["Товары"]
    assert discounted_sheet.max_row == 2
    assert discounted_sheet["D2"].value == "Товар со скидкой"
    assert discounted_sheet["J1"].value == "2026-06-10 price"
    assert discounted_sheet["K1"].value == "2026-06-10 discount_price"
    assert discounted_sheet["L1"].value == "2026-06-10 discount_percent"
    assert discounted_sheet["J2"].value == 110
    assert discounted_sheet["K2"].value == 88
    assert discounted_sheet["L2"].value == 20

    regular_output = ExportService(db_session).products_xlsx(
        category_id=good.id,
        has_discount=False,
        from_date=datetime(2026, 6, 10).date(),
        to_date=datetime(2026, 6, 10).date(),
    )
    regular_sheet = load_workbook(regular_output)["Товары"]
    assert regular_sheet.max_row == 2
    assert regular_sheet["D2"].value == "Товар без скидки"
    assert regular_sheet["J2"].value == 80


def test_products_export_period_uses_fullest_run_per_day(db_session) -> None:
    source, good, _ = seed_source_categories(db_session)
    first_run = ParserRun(
        source_id=source.id,
        status="success",
        started_at=datetime(2026, 6, 14, 9, tzinfo=timezone.utc),
        finished_at=datetime(2026, 6, 14, 9, 20, tzinfo=timezone.utc),
        total_categories=1,
        processed_categories=1,
        total_products=1,
        saved_products=1,
    )
    second_run = ParserRun(
        source_id=source.id,
        status="success",
        started_at=datetime(2026, 6, 14, 16, tzinfo=timezone.utc),
        finished_at=datetime(2026, 6, 14, 16, 20, tzinfo=timezone.utc),
        total_categories=1,
        processed_categories=1,
        total_products=2,
        saved_products=2,
    )
    db_session.add_all([first_run, second_run])
    db_session.flush()

    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=first_run.id,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="daily-a",
            sku="DAILY-A",
            name="Дневной товар А",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("100.00"),
            discount_price=None,
            discount_percent=None,
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/daily-a",
            is_available=True,
            raw_data={"id": "daily-a"},
        ),
        collected_at=datetime(2026, 6, 14, 9, 10, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=second_run.id,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="daily-a",
            sku="DAILY-A",
            name="Дневной товар А",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("120.00"),
            discount_price=None,
            discount_percent=None,
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/daily-a",
            is_available=True,
            raw_data={"id": "daily-a"},
        ),
        collected_at=datetime(2026, 6, 14, 16, 10, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=second_run.id,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="daily-b",
            sku="DAILY-B",
            name="Дневной товар Б",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("50.00"),
            discount_price=Decimal("45.00"),
            discount_percent=Decimal("10.00"),
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/daily-b",
            is_available=True,
            raw_data={"id": "daily-b"},
        ),
        collected_at=datetime(2026, 6, 14, 16, 12, tzinfo=timezone.utc),
    )
    db_session.commit()

    output = ExportService(db_session).products_xlsx(
        category_id=good.id,
        from_date=datetime(2026, 6, 14).date(),
        to_date=datetime(2026, 6, 14).date(),
    )
    sheet = load_workbook(output)["Товары"]
    rows_by_name = {sheet[f"D{row}"].value: row for row in range(2, sheet.max_row + 1)}

    assert sheet.max_row == 3
    assert sheet["J1"].value == "2026-06-14 price"
    assert sheet[f"J{rows_by_name['Дневной товар А']}"].value == 120
    assert sheet[f"J{rows_by_name['Дневной товар Б']}"].value == 50
    assert sheet[f"K{rows_by_name['Дневной товар Б']}"].value == 45


def test_products_export_filters_by_run_id(db_session) -> None:
    source, good, _ = seed_source_categories(db_session)
    first_run = ParserRun(
        source_id=source.id,
        status="success",
        started_at=datetime(2026, 6, 12, 9, tzinfo=timezone.utc),
        finished_at=datetime(2026, 6, 12, 9, 20, tzinfo=timezone.utc),
        total_categories=1,
        processed_categories=1,
        total_products=1,
        saved_products=1,
    )
    second_run = ParserRun(
        source_id=source.id,
        status="success",
        started_at=datetime(2026, 6, 13, 9, tzinfo=timezone.utc),
        finished_at=datetime(2026, 6, 13, 9, 20, tzinfo=timezone.utc),
        total_categories=1,
        processed_categories=1,
        total_products=1,
        saved_products=1,
    )
    db_session.add_all([first_run, second_run])
    db_session.flush()

    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=first_run.id,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="run-product",
            sku="RUN-1",
            name="Исторический товар",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("100.00"),
            discount_price=None,
            discount_percent=None,
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/run-product",
            is_available=True,
            raw_data={"id": "run-product"},
        ),
        collected_at=datetime(2026, 6, 12, 9, 10, tzinfo=timezone.utc),
    )
    SnapshotService(db_session).save_product_snapshot(
        source_id=source.id,
        category_id=good.id,
        run_id=second_run.id,
        parsed=ParsedProduct(
            source_code="globus",
            external_sku="run-product",
            sku="RUN-1",
            name="Исторический товар",
            unit="шт.",
            category_name=good.name,
            category_url=good.url,
            price=Decimal("140.00"),
            discount_price=None,
            discount_percent=None,
            image_url=None,
            product_url="https://globus-online.kg/ru-kg/good/run-product",
            is_available=True,
            raw_data={"id": "run-product"},
        ),
        collected_at=datetime(2026, 6, 13, 9, 10, tzinfo=timezone.utc),
    )
    db_session.commit()

    output = ExportService(db_session).products_xlsx(run_id=first_run.id)
    sheet = load_workbook(output)["Товары"]

    assert sheet.max_row == 2
    assert sheet["D2"].value == "Исторический товар"
    assert sheet["H2"].value == 100
    assert sheet["M2"].value == datetime(2026, 6, 12, 9, 10).replace(tzinfo=None)
